import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

headers = {
    "User-Agent": "Mozilla/5.0"
}

cantidad_paginas = 5

enlaces_boletines = []

for numero_pagina in range(1, cantidad_paginas + 1):

    if numero_pagina == 1:
        url = "https://conred.gob.gt/category/boletines-informativos/"
    else:
        url = f"https://conred.gob.gt/category/boletines-informativos/page/{numero_pagina}/"

    print("Revisando pagina:", url)

    respuesta = requests.get(url, headers=headers)

    if respuesta.status_code != 200:
        print("No se pudo acceder a esta pagina")
        continue

    sopa = BeautifulSoup(respuesta.text, "html.parser")

    articulos = sopa.find_all("article")

    for articulo in articulos:

        enlace_tag = articulo.find("a")

        if enlace_tag and enlace_tag.get("href"):

            enlace = enlace_tag.get("href")

            if enlace not in enlaces_boletines:
                enlaces_boletines.append(enlace)

print("Total de enlaces encontrados:", len(enlaces_boletines))

datos = []

for enlace in enlaces_boletines:

    print("Extrayendo informacion de:", enlace)

    respuesta = requests.get(enlace, headers=headers)

    if respuesta.status_code != 200:
        continue

    sopa = BeautifulSoup(respuesta.text, "html.parser")

    titulo_tag = sopa.find("h1")

    if titulo_tag:
        titulo = titulo_tag.get_text(strip=True)
    else:
        titulo = "Sin titulo"

    contenido_tag = sopa.find("article")

    if contenido_tag:
        contenido = contenido_tag.get_text(" ", strip=True)
    else:
        contenido = sopa.get_text(" ", strip=True)

    resumen = contenido[:300]

    texto_minuscula = contenido.lower()

    if "volcan" in texto_minuscula or "volcán" in texto_minuscula:
        clasificacion = "Actividad volcanica"

    elif "lluvia" in texto_minuscula or "inundacion" in texto_minuscula or "inundación" in texto_minuscula:
        clasificacion = "Lluvias o inundaciones"

    elif "sismo" in texto_minuscula or "terremoto" in texto_minuscula:
        clasificacion = "Actividad sismica"

    elif "incendio" in texto_minuscula:
        clasificacion = "Incendio"

    elif "prevencion" in texto_minuscula or "prevención" in texto_minuscula:
        clasificacion = "Prevencion"

    else:
        clasificacion = "General"

    datos.append([
        titulo,
        clasificacion,
        resumen,
        enlace
    ])

libro = Workbook()

hoja = libro.active
hoja.title = "Informacion CONRED"

encabezados = [
    "Titulo",
    "Clasificacion",
    "Resumen",
    "Enlace"
]

hoja.append(encabezados)

for fila in datos:
    hoja.append(fila)

borde_fino = Side(style="thin")

borde_completo = Border(
    left=borde_fino,
    right=borde_fino,
    top=borde_fino,
    bottom=borde_fino
)

for fila in hoja.iter_rows():

    for celda in fila:

        celda.border = borde_completo

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

for celda in hoja[1]:
    celda.font = Font(bold=True)

anchos = {
    "A": 55,
    "B": 25,
    "C": 75,
    "D": 70
}

for columna, ancho in anchos.items():
    hoja.column_dimensions[columna].width = ancho

for numero_fila in range(1, hoja.max_row + 1):
    hoja.row_dimensions[numero_fila].height = 100

hoja.row_dimensions[1].height = 30

libro.save("conred_informacion.xlsx")

print("Datos extraidos correctamente")
print("Archivo generado: conred_informacion.xlsx")